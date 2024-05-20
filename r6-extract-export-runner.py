from flask import Flask
import subprocess
import os
import json
import csv
import logging
from openpyxl import Workbook

# Set up logging
logging.basicConfig(level=logging.INFO)

app = Flask(__name__)

def download_s3_file(bucket, file, local_path):
    s3_file_path = os.path.join(bucket, file)
    print(f"Downloading {s3_file_path} to {local_path}")
    logging.info(f"Downloading {s3_file_path} to {local_path}")
    local_file_name = os.path.basename(file)
    local_file_path = os.path.join(local_path, local_file_name)
    cmd = f"aws s3 cp {s3_file_path} {local_file_path}"
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"Failed to download {file}: {result.stderr}")
        logging.error(f"Failed to download {file}: {result.stderr}")
    else:
        print(f"Successfully downloaded {file} to {local_file_path}")
        logging.info(f"Successfully downloaded {file} to {local_file_path}")
        process_json_to_xlsx(local_file_path)

def process_json_to_xlsx(json_path):
    with open(json_path, 'r') as json_file:
        data = json.load(json_file)

    xlsx_path = json_path.replace('.json', '.xlsx')
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    ws_stats = wb.create_sheet("Stats")
    ws_per_round = wb.create_sheet("Per Round")
    ws_per_match = wb.create_sheet("Per Match")
    
    # Write headers to the Stats sheet
    ws_stats.append(['TEAMS'])  # Column header

    # If there are rounds, process the first one to get team names
    first_match = data['rounds'][0] if data['rounds'] else None
    if first_match:
        teams = first_match.get('teams', [])
        for team in teams:
            ws_stats.append([team.get('name', 'No Team Name')])  # Add team names

    # Save the workbook
    wb.save(xlsx_path)
    print(f"Excel file saved to {xlsx_path}")


def list_s3_files(bucket):
    cmd = f"aws s3 ls {bucket} --recursive"
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"Error listing S3 bucket contents: {result.stderr}")
        logging.error(f"Error listing S3 bucket contents: {result.stderr}")
        return []
    return [line.split()[-1] for line in result.stdout.split('\n') if line.strip()]

@app.route('/api/runner', methods=['POST'])
def runner():
    data_directory = "/app/data"
    os.makedirs(data_directory, exist_ok=True)
    intermediate_data_bucket = "s3://tlmrisserver/intermediate-data"
    post_exported_data_bucket = "s3://tlmrisserver/post-exported-data"

    intermediate_files = list_s3_files(intermediate_data_bucket)
    post_exported_files = list_s3_files(post_exported_data_bucket)

    intermediate_files_set = set(intermediate_files)
    post_exported_files_set = set(post_exported_files)

    unique_files = intermediate_files_set - post_exported_files_set

    for file in unique_files:
        download_s3_file("s3://tlmrisserver/", file, data_directory)

    return "File download and processing complete!"

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5002)
